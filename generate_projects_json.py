#!/usr/bin/env python3
"""
generate_projects_json.py
Run by GitHub Action after sync_data.yml to produce projects.json.gz
Pre-aggregates 825K DN dump rows → ~32K project-level rows (~1.2 MB gz vs 14.7 MB)

Metering is calculated per-project using the backend rate table formula:
  metering = NM_rate(city, phase) + GM_rate(city, phase) + DN_dump_metering_items
No manual monthly totals needed — fully automatic for any month.

COGS uses the ERP Categorization file as the PRIMARY source for item_code → category
mapping, falling back to the raw DN item_category only when the item_code is not in
the ERP file.  Safety Lifeline and Civil Work are excluded from COGS.

Usage:  python3 generate_projects_json.py
Input:  data.csv.gz  (same directory)
        erp_categorization.csv (same directory — optional, uses embedded map if missing)
Output: projects.json.gz (same directory)
"""
import gzip, csv, io, json, os, re, sys
from collections import defaultdict
from datetime import datetime

# ── Cohort assignment (QCD-based) ────────────────────────────────────────────
# Cohorts are assigned by QCD (Quote Completion Date), NOT installation date.
# QCD is loaded from booking_dump.csv (exported from OMS import sheet).
# Pricing cohort definitions come from pricing cohorts.xlsx.
# GZ offers: 'GoodZero', 'GoodZero Pro', 'GoodZero Uno', 'GoodZero+'
# Non-GZ offers: 'Regular', 'regular', 'SSE Blue', '' (blank), etc.

GZ_OFFERS_SET = {'GoodZero', 'GoodZero Pro', 'GoodZero Uno', 'GoodZero+'}

# ── Configuration ────────────────────────────────────────────────────────────
CIVL_TO_ELEC   = {'CIVL-0012','CIVL-0013','CIVL-0014','CIVL-0015','CIVL-0016'}
METERING_REMAP = {'ACDB-2449-EATON'}
DONGLE_PFX     = {'DALO','DALA'}

# Categories that count towards COGS
COGS_CATS = {
    'Module','Inverter','Prefab MMS','Cables','I&C KIT','Conduit Pipe',
    'Earthing & LA','Junction Box','Tin Shed MMS','Safety','I&C Accessories',
    'Welded MMS','SS NBW','Electrical BoS','Data Logger','Metering','Welcome Kit and Board','Ladder'
}

# Categories explicitly EXCLUDED from COGS (even if present in DN data)
EXCLUDE_CATS = {'Safety Lifeline', 'Civil Work', 'Civil work'}

# ── Backend Metering Rate Tables ─────────────────────────────────────────────
# Formula: metering = NM(city, inv_phase) + GM(city, sanction_phase) + DN_dump_extras
# NM = Net Meter rate (keyed on Inverter Phase — detected from DN inverter item name)
# GM = Generation Meter rate (keyed on Sanction Phase = Phase Connection from data.csv)
# Tuple format: (single_phase, three_phase)
# Source: Backend rate matrix (GMB_GMP_GMI ERP Categorization)

NM_RATES = {
    # MH clusters — Net Meter included in discom, rate = 0
    'Pune':        (0, 0),
    'Nashik':      (0, 0),
    'Nagpur':      (0, 0),
    'Aurangabad':  (0, 0),
    'Jalgaon':     (0, 0),
    'Ahmednagar':  (0, 0),
    'Latur':       (0, 0),
    'Kolhapur':    (0, 0),
    'Mumbai':      (0, 0),
    'Amravati':    (0, 0),
    'Solapur':     (0, 0),
    # MP clusters
    'Bhopal':      (2841, 4617),
    'Indore':      (6800, 9050),
    'Jabalpur':    (9785, 14050),
    'Gwalior':     (2841, 4617),
    # South
    'Bengaluru':   (3250, 6376),
    'Hyderabad':   (0, 0),
    # Gujarat
    'Ahmedabad':   (0, 0),
    'Surat':       (0, 0),
    'Baroda':      (0, 0),
    # Rajasthan
    'Jaipur':      (3550, 6650),
    'Ajmer':       (3550, 6650),
    'Kota':        (3550, 6650),
    # UP / North
    'Lucknow':     (1350, 4350),
    'Kanpur':      (1350, 4350),
    'Varanasi':    (1350, 4350),
    'Noida':       (1350, 4350),
    'NCR':         (0, 0),
    'Meerut':      (1350, 4350),
    'Bareilly':    (1350, 4350),
    # South (others)
    'Kochi':       (3250, 6376),
    'Chennai':     (2763, 5011),
    'Agra':        (1350, 4350),
    'Coimbatore':  (2763, 5011),
    # Additional mapped cities
    'Raipur':      (0, 0),
    'Mysuru':      (3250, 6376),
    'Warangal':    (0, 0),
    'Gurgaon':     (0, 0),
    'Delhi NCR':   (0, 0),
    'Ghaziabad':   (1350, 4350),
    'Vijayawada':  (0, 0),
}

GM_RATES = {
    # MH clusters — Generation Meter procured by SSE
    'Pune':        (1260, 2620),
    'Nashik':      (1260, 2620),
    'Nagpur':      (1260, 2620),
    'Aurangabad':  (1260, 2620),
    'Jalgaon':     (1260, 2620),
    'Ahmednagar':  (1260, 2620),
    'Latur':       (1260, 2620),
    'Kolhapur':    (1260, 2620),
    'Mumbai':      (1260, 2620),
    'Amravati':    (1260, 2620),
    'Solapur':     (1260, 2620),
    # MP clusters — no Gen Meter
    'Bhopal':      (0, 0),
    'Indore':      (0, 0),
    'Jabalpur':    (0, 0),
    'Gwalior':     (0, 0),
    # South
    'Bengaluru':   (0, 0),
    'Hyderabad':   (0, 0),
    # Gujarat
    'Ahmedabad':   (0, 0),
    'Surat':       (0, 0),
    'Baroda':      (0, 0),
    # Rajasthan — Gen Meter applies
    'Jaipur':      (3050, 5650),
    'Ajmer':       (3050, 5650),
    'Kota':        (3050, 5650),
    # UP / North — no Gen Meter
    'Lucknow':     (0, 0),
    'Kanpur':      (0, 0),
    'Varanasi':    (0, 0),
    'Noida':       (0, 0),
    'NCR':         (0, 0),
    'Meerut':      (0, 0),
    'Bareilly':    (0, 0),
    # South (others)
    'Kochi':       (0, 0),
    'Chennai':     (0, 0),
    'Agra':        (0, 0),
    'Coimbatore':  (0, 0),
    # Additional mapped cities
    'Raipur':      (0, 0),
    'Mysuru':      (0, 0),
    'Warangal':    (0, 0),
    'Gurgaon':     (0, 0),
    'Delhi NCR':   (0, 0),
    'Ghaziabad':   (0, 0),
}

def detect_inverter_phase(inv_item_name):
    """Detect inverter phase from DN item_name.
    NM rate uses Inverter Phase (not Sanction Phase).
    Returns 'Single Phase' or 'Three Phase'."""
    n = str(inv_item_name)
    # Skip batteries — they aren't the main inverter
    if 'Battery' in n and 'Hybrid' in n:
        return None  # signal to skip this item
    if '3 Phase' in n or '3-Phase' in n or 'Three Phase' in n:
        return 'Three Phase'
    if '1 Phase' in n or '1-Phase' in n or 'Single Phase' in n:
        return 'Single Phase'
    # Enphase microinverters are always single phase
    if 'ENPHASE' in n.upper() or 'Micro' in n.lower():
        return 'Single Phase'
    # Default: single phase
    return 'Single Phase'

def calc_metering_backend(city, inv_phase, sanction_phase):
    """Calculate backend metering = NM_rate(city, phase) + GM_rate(city, sanction_phase).
    NM uses Sanction Phase (Phase Connection) as primary, with inverter phase as fallback.
    GM uses Sanction Phase (= Phase Connection from data.csv).
    Using sanction_phase for NM too gives better accuracy since detect_inverter_phase
    can default single-phase for 3-phase projects when item name pattern is ambiguous."""
    # NM lookup — use sanction_phase as primary (more reliable), inv_phase as fallback
    phase_for_nm = sanction_phase if sanction_phase else inv_phase
    nm_idx = 0 if (not phase_for_nm or 'single' in phase_for_nm.lower()) else 1
    nm = NM_RATES.get(city, (0, 0))
    # GM lookup — keyed on sanction/connection phase
    gm_idx = 0 if (not sanction_phase or 'single' in sanction_phase.lower()) else 1
    gm = GM_RATES.get(city, (0, 0))
    return nm[nm_idx] + gm[gm_idx]

def is_metering_dn_item(item_name):
    """Check if a DN dump item matches the metering SUMIFS patterns"""
    if 'Communication Modem' in item_name and 'Optical Cable' in item_name:
        return True
    if 'FRP Meter Box' in item_name:
        return True
    if 'Meter Box' in item_name and '400x300x150' in item_name and 'SPARK' in item_name:
        return True
    return False

# ── Cell Name → City/State Lookup ────────────────────────────────────────────
CELL_CITY_STATE = {
    'Aurangabad Expansion':{'c':'Aurangabad','s':'MH East'},
    'Bangalore Royal Challengers':{'c':'Bengaluru','s':'Karnataka'},
    'Bangalore Royal Challengers**':{'c':'Bengaluru','s':'Karnataka'},
    'Bangalore Royal Challengers 2':{'c':'Bengaluru','s':'Karnataka'},
    'Bengaluru Royal Challengers':{'c':'Bengaluru','s':'Karnataka'},
    'Bengaluru Royal Challengers 2':{'c':'Bengaluru','s':'Karnataka'},
    'Baroda Blasters':{'c':'Baroda','s':'Gujrat'},
    'Baroda Smashers':{'c':'Baroda','s':'Gujrat'},
    'Bhopal Strikers':{'c':'Bhopal','s':'Madhya Pradesh'},
    'Bhopal Strikers 2':{'c':'Bhopal','s':'Madhya Pradesh'},
    'Bhopal Strikers 3':{'c':'Bhopal','s':'Madhya Pradesh'},
    'Bhopal Strikers 4':{'c':'Bhopal','s':'Madhya Pradesh'},
    'Bhopal Strikers 5':{'c':'Bhopal','s':'Madhya Pradesh'},
    'Bhopal Strikers 6':{'c':'Bhopal','s':'Madhya Pradesh'},
    'Bhopal Strikers**':{'c':'Bhopal','s':'Madhya Pradesh'},
    'Delhi Dashers 2':{'c':'Gurgaon','s':'Delhi'},
    'Delhi Dashers 3':{'c':'Delhi NCR','s':'Delhi'},
    'Delhi Dashers 4':{'c':'Ghaziabad','s':'Delhi'},
    'Delhi Dashers 5':{'c':'Delhi NCR','s':'Delhi'},
    'Delhi Dashers 6':{'c':'Delhi NCR','s':'Delhi'},
    'Gujrat Gladiators':{'c':'Ahmedabad','s':'Gujrat'},
    'Gujrat Gladiators 2':{'c':'Ahmedabad','s':'Gujrat'},
    'Gujarat Gladiators':{'c':'Ahmedabad','s':'Gujrat'},
    'Gujarat Gladiators 2':{'c':'Ahmedabad','s':'Gujrat'},
    'Ahmedabad Gladiators':{'c':'Ahmedabad','s':'Gujrat'},
    'Ahmedabad Gladiators 2':{'c':'Ahmedabad','s':'Gujrat'},
    'Gwalior Groundbreakers':{'c':'Gwalior','s':'Madhya Pradesh'},
    'Gwalior Groundbreakers 2':{'c':'Gwalior','s':'Madhya Pradesh'},
    'Gwalior Groundbreakers 3':{'c':'Gwalior','s':'Madhya Pradesh'},
    'Gwalior Groundbreakers 4':{'c':'Gwalior','s':'Madhya Pradesh'},
    'Gwalior Groundbreakers 5':{'c':'Gwalior','s':'Madhya Pradesh'},
    'Speed Order Gwalior 5':{'c':'Gwalior','s':'Madhya Pradesh'},
    'Indore Immortals':{'c':'Indore','s':'Madhya Pradesh'},
    'Indore Immortals 2':{'c':'Indore','s':'Madhya Pradesh'},
    'Indore Immortals 3':{'c':'Indore','s':'Madhya Pradesh'},
    'Indore Immortals 4':{'c':'Indore','s':'Madhya Pradesh'},
    'Indore Immortals 5':{'c':'Indore','s':'Madhya Pradesh'},
    'Indore Immortals 6':{'c':'Indore','s':'Madhya Pradesh'},
    'Indore Immortals 7':{'c':'Indore','s':'Madhya Pradesh'},
    'Jabalpur Champions':{'c':'Jabalpur','s':'Madhya Pradesh'},
    'Jabalpur Champions 2':{'c':'Jabalpur','s':'Madhya Pradesh'},
    'Jabalpur Champions 3':{'c':'Jabalpur','s':'Madhya Pradesh'},
    'Jabalpur Champions 4':{'c':'Jabalpur','s':'Madhya Pradesh'},
    'Jabalpur Champions 5':{'c':'Jabalpur','s':'Madhya Pradesh'},
    'Jalgaon Expansion':{'c':'Jalgaon','s':'MH East'},
    'Jalgaon Expansion 2':{'c':'Jalgaon','s':'MH East'},
    'Kolhapur Kings':{'c':'Kolhapur','s':'MH West'},
    'Lucknow Lions':{'c':'Lucknow','s':'Uttar Pradesh'},
    'Lucknow Lions 2':{'c':'Lucknow','s':'Uttar Pradesh'},
    'Lucknow Lions 3':{'c':'Lucknow','s':'Uttar Pradesh'},
    'Lucknow Lions 4':{'c':'Lucknow','s':'Uttar Pradesh'},
    'Speed Order Lucknow 4':{'c':'Lucknow','s':'Uttar Pradesh'},
    'Speed Order Lucknow 5':{'c':'Lucknow','s':'Uttar Pradesh'},
    'Noida Knight Riders':{'c':'Noida','s':'Uttar Pradesh'},
    'Kanpur Tigers':{'c':'Kanpur','s':'Uttar Pradesh'},
    'Kanpur Tigers 2':{'c':'Kanpur','s':'Uttar Pradesh'},
    'Kanpur Tigers 3':{'c':'Kanpur','s':'Uttar Pradesh'},
    'Varanasi Warriors':{'c':'Varanasi','s':'Uttar Pradesh'},
    'Agra Knight Riders':{'c':'Agra','s':'Uttar Pradesh'},
    'Nagpur Daredevils':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 2':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 3':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 4':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 5':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 6':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 7':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 8':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 9':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 10':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 13':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 14':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils 15':{'c':'Nagpur','s':'MH East'},
    'Nagpur Daredevils Temp':{'c':'Nagpur','s':'MH East'},
    'Amravati Riders':{'c':'Amravati','s':'MH East'},
    'Amravati Riders 3':{'c':'Amravati','s':'MH East'},
    'Nashik Finishers':{'c':'Nashik','s':'MH West'},
    'Nashik Finishers 2':{'c':'Nashik','s':'MH West'},
    'Nashik Finishers 3':{'c':'Nashik','s':'MH West'},
    'Nashik Finishers 5':{'c':'Nashik','s':'MH West'},
    'Pune Squadrons':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 2':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 3':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 4':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 5':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 6':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 7':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 8':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 9':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 10':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 11':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 12':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 13':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 14':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 15':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 16':{'c':'Pune','s':'MH West'},
    'Pune Squadrons 17':{'c':'Pune','s':'MH West'},
    'Pune Squadrons Temp':{'c':'Pune','s':'MH West'},
    'Pune Squadrons**':{'c':'Pune','s':'MH West'},
    'Ahilyanagar Regiments':{'c':'Pune','s':'MH West'},
    'Speed Order Ahilyanagar 1':{'c':'Pune','s':'MH West'},
    'Speed Order Pune 11':{'c':'Pune','s':'MH West'},
    'Solapur Super Kings':{'c':'Solapur','s':'MH West'},
    'Surat Expansion':{'c':'Surat','s':'Gujrat'},
    'Surat Expansion 2':{'c':'Surat','s':'Gujrat'},
    'Jaipur Titans':{'c':'Jaipur','s':'Rajasthan'},
    'Speed Order Jaipur 2':{'c':'Jaipur','s':'Rajasthan'},
    'Kota Knights':{'c':'Kota','s':'Rajasthan'},
    'Ajmer Aces':{'c':'Ajmer','s':'Rajasthan'},
    'Ajmer Aces 2':{'c':'Ajmer','s':'Rajasthan'},
    'Telangana Tuskers':{'c':'Hyderabad','s':'Telangana'},
    'Telangana Tuskers 2':{'c':'Hyderabad','s':'Telangana'},
    'Telangana Tuskers 3':{'c':'Hyderabad','s':'Telangana'},
    'Hyderabad Tuskers':{'c':'Hyderabad','s':'Telangana'},
    'Hyderabad Tuskers 2':{'c':'Hyderabad','s':'Telangana'},
    'Hyderabad Tuskers 3':{'c':'Hyderabad','s':'Telangana'},
    'Warangal Waveriders':{'c':'Warangal','s':'Telangana'},
    'Vijayawada Strikers':{'c':'Vijayawada','s':'Andhra Pradesh'},
    'Kochi Crushers':{'c':'Kochi','s':'Kerala'},
    'Raipur Royals':{'c':'Raipur','s':'Chhattisgarh'},
    'Chennai Super Kings':{'c':'Chennai','s':'Tamil Nadu'},
    'Chennai Super Kings 2':{'c':'Chennai','s':'Tamil Nadu'},
    'Speed Order Chennai 3':{'c':'Chennai','s':'Tamil Nadu'},
    'Coimbatore Kovai Kings':{'c':'Coimbatore','s':'Tamil Nadu'},
    'Mysuru Mavericks':{'c':'Mysuru','s':'Karnataka'},
    'Speed Order Gurgaon':{'c':'Gurgaon','s':'Delhi'},
    'Latur Expansion':{'c':'Latur','s':'MH East'},
    'Ahmednagar Expansion':{'c':'Ahmednagar','s':'MH West'},
}

def detect_inverter_type(item_name):
    """Extract inverter type from DN item name for dashboard type-level analysis.
    Returns simplified type string like '3 kW', '5 kW 3 Phase', '5 kW Hybrid', 'Enphase'."""
    n = str(item_name)
    # Skip batteries
    if 'Battery' in n and 'Hybrid' in n:
        return None
    # Enphase microinverters
    if 'ENPHASE' in n.upper() or 'Micro' in n.lower():
        return 'Enphase'
    # Extract kW rating
    m = re.search(r'(\d+\.?\d*)\s*[kK][wW]', n)
    if not m:
        return 'Other'
    kw = m.group(1)
    # Normalize: remove trailing .0
    try:
        kw_f = float(kw)
        kw = str(int(kw_f)) if kw_f == int(kw_f) else str(kw_f)
    except: pass
    # Check for Hybrid
    if 'Hybrid' in n:
        return f'{kw} kW Hybrid'
    # Check for phase
    if '3 Phase' in n or '3-Phase' in n or 'Three Phase' in n:
        return f'{kw} kW 3 Phase'
    return f'{kw} kW'

MON_MAP = {'jan':0,'feb':1,'mar':2,'apr':3,'may':4,'jun':5,'jul':6,'aug':7,'sep':8,'oct':9,'nov':10,'dec':11}

def parse_date(v):
    if not v: return None
    parts = v.strip().split('-')
    if len(parts) == 3:
        try:
            day = int(parts[0])
            mon = MON_MAP.get(parts[1].lower()[:3])
            yr  = int(parts[2]); yr = 2000+yr if yr < 100 else yr
            if mon is not None: return datetime(yr, mon+1, day)
        except: pass
    try: return datetime.strptime(v.strip(), '%Y-%m-%d')
    except: pass
    try: return datetime.strptime(v.strip(), '%d/%m/%Y')
    except: pass
    try: return datetime.strptime(v.strip(), '%Y/%m/%d')
    except: pass
    return None


# ── Pricing Cohort Loader ─────────────────────────────────────────────────────
# Loads cohort date ranges from pricing cohorts.xlsx.
# Returns two sorted lists: [(start, end_or_None, name), ...] for GZ and Non-GZ.
# 'end' is inclusive; None means open-ended (current cohort).

def load_pricing_cohorts(filepath='pricing cohorts.xlsx'):
    """Load GZ and Non-GZ cohort date ranges from pricing cohorts.xlsx."""
    gz_cohorts   = []   # list of (start_date, end_date_or_None, cohort_name)
    non_gz_cohorts = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb['Sheet1']
        rows = list(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True))  # skip 2 header rows
        for r in rows:
            gz_start, gz_end, gz_name = r[0], r[1], r[2]
            ngz_start, ngz_end, ngz_name = r[3], r[4], r[5]
            if gz_name and gz_start:
                s = gz_start if isinstance(gz_start, datetime) else parse_date(str(gz_start))
                e = gz_end   if isinstance(gz_end,   datetime) else (parse_date(str(gz_end)) if gz_end else None)
                if s: gz_cohorts.append((s, e, str(gz_name).strip()))
            if ngz_name:
                s = ngz_start if isinstance(ngz_start, datetime) else (parse_date(str(ngz_start)) if ngz_start else None)
                e = ngz_end   if isinstance(ngz_end,   datetime) else (parse_date(str(ngz_end))   if ngz_end   else None)
                non_gz_cohorts.append((s, e, str(ngz_name).strip()))
        gz_cohorts.sort(key=lambda x: x[0])
        non_gz_cohorts.sort(key=lambda x: (x[0] or datetime.min))
        print(f"  Loaded {len(gz_cohorts)} GZ cohorts, {len(non_gz_cohorts)} Non-GZ cohorts from {filepath}")
    except ImportError:
        print("  ⚠ openpyxl not installed — cohort assignment disabled. Run: pip install openpyxl")
    except FileNotFoundError:
        print(f"  ⚠ {filepath} not found — cohort assignment disabled.")
    except Exception as ex:
        print(f"  ⚠ Could not load pricing cohorts: {ex}")
    return gz_cohorts, non_gz_cohorts

GZ_COHORTS, NON_GZ_COHORTS = load_pricing_cohorts()


def assign_cohort(qcd_date, offer_type):
    """Assign a pricing cohort name based on QCD date and offer type.

    Args:
        qcd_date: datetime object (Quote Completion Date from booking dump)
        offer_type: str — raw offer from OMS/DN ('GoodZero', 'Regular', 'SSE Blue', etc.)

    Returns:
        str: cohort name, or '' if QCD date is missing or no cohort matches.

    Boundary rule: end date is EXCLUSIVE (cohort covers start <= qcd < end).
    This means when two cohorts share the same date (e.g., 9th Apr ends, 15th Apr starts,
    both listing Apr 15), a QCD of Apr 15 falls into the NEWER cohort ("15th Apr Onwards").
    This matches the business logic: the cohort name says "15th Apr Onwards" — Apr 15 IS
    the start of that cohort, not the last day of the previous one.
    """
    if not qcd_date:
        return ''
    is_gz = offer_type.strip().replace('GoodZero+','GoodZero') in GZ_OFFERS_SET
    cohorts = GZ_COHORTS if is_gz else NON_GZ_COHORTS
    # Iterate in REVERSE so newer (later-starting) cohorts take priority at shared boundaries
    for (start, end, name) in reversed(cohorts):
        if start is None:
            # open-start bucket (e.g. "Before Amit's pricing") — end is inclusive here
            if end and qcd_date <= end:
                return name
        else:
            # end is exclusive: cohort covers [start, end)
            if qcd_date >= start and (end is None or qcd_date < end):
                return name
    return ''


# ── Booking Dump Loader (QCD dates) ──────────────────────────────────────────
# PRIMARY:  Fetches live from Google Sheets every time the script runs.
#           Sheet is public (anyone with link) — no auth required.
#           URL = OMS Import Sheet → "Booking Dump" tab (gid=628408580)
# FALLBACK: If the fetch fails (no internet, quota, etc.), reads local
#           booking_dump.csv cached from the last successful fetch.
#
# Output: dict  { sse_id: {'qcd': datetime_or_None, 'offer': str} }

BOOKING_DUMP_URL = (
    'https://docs.google.com/spreadsheets/d/'
    '1NmE-MH9NyLFcbX1JH--j3yqT32sahvJGj82uFK6l9CY'
    '/gviz/tq?tqx=out:csv&gid=628408580'
)
BOOKING_DUMP_CACHE = 'booking_dump.csv'   # local cache written after every successful fetch


def _detect_booking_cols(headers):
    """Auto-detect SSE ID, QCD date, and Offer Type column names.
    Handles trailing spaces, slashes, mixed case (e.g. 'QCD / LQUD ').
    Returns (sse_col, qcd_col, offer_col) — offer_col may be None."""
    sse_col = qcd_col = offer_col = None
    for h in headers:
        hl = h.strip().lower().replace(' ','').replace('_','').replace('-','').replace('/','')
        if not sse_col and hl in ('sseid','ssid','projectid','projid','sseno','sse'):
            sse_col = h
        if not qcd_col and hl in ('qcd','qcdlqud','lqud','qcddate','quotecompletiondate',
                                   'quoteclosuredate','quotationdate','closuredate',
                                   'quotedate','qcdate','quotationcreationdate'):
            qcd_col = h
        if not offer_col and hl in ('offertype','offeringtype','offer','product',
                                    'producttype','schemetype'):
            offer_col = h
    # Broader fallbacks
    if not sse_col:
        for h in headers:
            if 'sse' in h.lower(): sse_col = h; break
    if not qcd_col:
        for h in headers:
            hl = h.strip().lower()
            if 'qcd' in hl or 'lqud' in hl: qcd_col = h; break
    return sse_col, qcd_col, offer_col


def _parse_booking_rows(rows_raw):
    """Convert raw DictReader rows into qcd_map.
    Strips trailing spaces from all keys and values."""
    qcd_map = {}
    if not rows_raw:
        return qcd_map
    headers = list(rows_raw[0].keys())
    sse_col, qcd_col, offer_col = _detect_booking_cols(headers)
    if not sse_col:
        print(f"    ⚠ SSE ID column not found. Headers: {headers[:10]}")
        return qcd_map
    if not qcd_col:
        print(f"    ⚠ QCD date column not found. Headers: {headers[:15]}")
        print(f"       Expected: 'QCD', 'QCD / LQUD', 'LQUD', 'QCD Date', etc.")
        return qcd_map
    loaded = 0
    for row in rows_raw:
        sse = row.get(sse_col, '').strip()
        if not sse:
            continue
        qcd_raw = row.get(qcd_col, '')
        qcd_dt = None
        if isinstance(qcd_raw, datetime):
            qcd_dt = qcd_raw
        elif qcd_raw:
            qcd_dt = parse_date(str(qcd_raw).strip())
        offer_raw = row.get(offer_col, '').strip() if offer_col else ''
        qcd_map[sse] = {'qcd': qcd_dt, 'offer': offer_raw}
        loaded += 1
    valid = sum(1 for v in qcd_map.values() if v['qcd'])
    print(f"    Parsed {loaded:,} rows → {valid:,} with valid QCD dates")
    return qcd_map


def load_booking_dump():
    """Fetch QCD dates from Google Sheets (live) with local CSV fallback.

    Flow:
      1. Fetch live CSV from BOOKING_DUMP_URL (public Google Sheet).
      2. On success → parse, save to BOOKING_DUMP_CACHE, return map.
      3. On any failure → warn, read BOOKING_DUMP_CACHE instead.
      4. If cache also missing → return empty map (cohort field stays blank).
    """
    import urllib.request, io

    qcd_map = {}

    # ── Step 1: Try live fetch ────────────────────────────────────────────────
    print(f"  Fetching booking dump from Google Sheets...")
    fetched_csv = None
    try:
        req = urllib.request.Request(
            BOOKING_DUMP_URL,
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            fetched_csv = resp.read().decode('utf-8', errors='replace')
        print(f"    ✅ Fetched {len(fetched_csv):,} chars from Google Sheets")
    except Exception as ex:
        print(f"    ⚠ Live fetch failed: {ex}")

    # ── Step 2: Parse fetched CSV if we got it ────────────────────────────────
    if fetched_csv:
        try:
            reader = csv.DictReader(io.StringIO(fetched_csv))
            rows_raw = list(reader)
            if rows_raw:
                qcd_map = _parse_booking_rows(rows_raw)
                if qcd_map:
                    # Save as local cache for next fallback
                    try:
                        with open(BOOKING_DUMP_CACHE, 'w', encoding='utf-8', newline='') as f:
                            writer = csv.DictWriter(f, fieldnames=list(rows_raw[0].keys()))
                            writer.writeheader()
                            writer.writerows(rows_raw)
                        print(f"    💾 Cached {len(rows_raw):,} rows → {BOOKING_DUMP_CACHE}")
                    except Exception as cache_ex:
                        print(f"    ⚠ Could not save cache: {cache_ex}")
                    return qcd_map
            print("    ⚠ Fetched CSV was empty — falling back to cache")
        except Exception as parse_ex:
            print(f"    ⚠ Could not parse fetched CSV: {parse_ex} — falling back to cache")

    # ── Step 3: Fallback to local cache ──────────────────────────────────────
    if os.path.isfile(BOOKING_DUMP_CACHE):
        print(f"  Reading cached booking dump from {BOOKING_DUMP_CACHE}...")
        try:
            with open(BOOKING_DUMP_CACHE, 'r', encoding='utf-8', errors='replace') as f:
                rows_raw = list(csv.DictReader(f))
            qcd_map = _parse_booking_rows(rows_raw)
            if qcd_map:
                import os as _os
                mtime = datetime.fromtimestamp(_os.path.getmtime(BOOKING_DUMP_CACHE))
                print(f"    ⚠ Using cached data (last updated: {mtime.strftime('%Y-%m-%d %H:%M')})")
                return qcd_map
        except Exception as ex:
            print(f"    ⚠ Could not read cache: {ex}")

    # ── Step 4: Nothing worked ────────────────────────────────────────────────
    print("  ⚠ Booking dump unavailable — cohort field will be blank for all projects.")
    print(f"    Fix: ensure the Google Sheet is public OR place {BOOKING_DUMP_CACHE} locally.")
    return qcd_map


print("\nLoading booking dump (QCD dates)...")
QCD_MAP = load_booking_dump()
print(f"  QCD map size: {len(QCD_MAP):,} projects")


# ── ERP Categorization ───────────────────────────────────────────────────────
# Load from erp_categorization.csv if available; this file maps item_code → category
# and overrides the raw DN item_category field.
# Expected CSV columns: item_code, item_category  (or Item Code, Category)
# The file is the authoritative source — raw DN categories are often empty/wrong.

ERP_CAT_MAP = {}

def load_erp_categorization():
    """Load ERP categorization CSV. Try multiple possible filenames."""
    global ERP_CAT_MAP
    candidates = [
        'erp_categorization.csv',
        'GMB_GMP_GMI_Mar_26_ERP_Categorization.csv',
        'GMB_GMP_GMI_ERP_Categorization.csv',
        'erp_cat.csv',
    ]
    for fname in candidates:
        if os.path.isfile(fname):
            print(f"Loading ERP categorization from {fname}...")
            with open(fname, 'r', encoding='utf-8', errors='replace') as f:
                reader = csv.DictReader(f)
                headers = reader.fieldnames
                # Detect column names (flexible matching)
                code_col = None
                cat_col  = None
                for h in headers:
                    hl = h.strip().lower()
                    if hl in ('item_code', 'item code', 'itemcode', 'code'):
                        code_col = h
                    if hl in ('item_category', 'item category', 'category', 'itemcategory'):
                        cat_col = h
                if not code_col or not cat_col:
                    print(f"  ⚠ Could not find item_code/category columns in {fname}")
                    print(f"    Headers found: {headers}")
                    continue
                count = 0
                for row in reader:
                    ic = row[code_col].strip()
                    ca = row[cat_col].strip()
                    if ic and ca:
                        ERP_CAT_MAP[ic] = ca
                        count += 1
                print(f"  Loaded {count:,} item_code → category mappings")
                return True
    print("  ⚠ No ERP categorization file found — using raw DN categories only")
    print("    (Place erp_categorization.csv in the same directory for accurate COGS)")
    return False

erp_loaded = load_erp_categorization()


def resolve_cat(item_code, raw_cat, item_subcategory=''):
    """Resolve the COGS category for a DN line item.
    Priority: hardcoded overrides > ERP file > raw DN category."""
    pfx = item_code[:4].upper()

    # 1. Hardcoded exclusions/remaps (highest priority)
    if pfx in DONGLE_PFX:                     return 'EXCLUDE'
    if item_code in CIVL_TO_ELEC:             return 'Electrical BoS'
    if item_code in METERING_REMAP:           return 'Metering'

    # 2. ERP categorization file override (primary source)
    erp = ERP_CAT_MAP.get(item_code)
    if erp:
        # Apply same transforms as raw_cat
        if erp in EXCLUDE_CATS:               return 'EXCLUDE'
        if erp == 'Fixtures and Tools':
            if item_subcategory in ('Aluminium Ladder', 'Ladder'):
                return 'Ladder'
            return 'Welcome Kit and Board'
        return erp.strip()

    # 3. Fallback to raw DN category
    if raw_cat in EXCLUDE_CATS:               return 'EXCLUDE'
    if raw_cat == 'Fixtures and Tools':
        if item_subcategory in ('Aluminium Ladder', 'Ladder'):
            return 'Ladder'
        return 'Welcome Kit and Board'
    if not raw_cat and item_code.startswith('INVS'): return 'Inverter'
    return raw_cat.strip()


CAT_KEY = {
    'Module':'mod','Inverter':'inv','Prefab MMS':'prf','Cables':'cab','I&C KIT':'ick',
    'Conduit Pipe':'con','Earthing & LA':'ear','Junction Box':'jbx','Tin Shed MMS':'tsh',
    'Safety':'saf','I&C Accessories':'ica','Welded MMS':'wel','SS NBW':'ssn',
    'Electrical BoS':'ebo','Data Logger':'dlg','Metering':'mtr','Welcome Kit and Board':'wkt',
    'Ladder':'lad',
}

# ── Name shortening for dashboard display ─────────────────────────────────────
def shorten_mms_item_name(name, subcat=''):
    """Shorten MMS item names for cleaner dashboard display.
    Column items: 'Column 2P 6FT Back Medium Gen 2 (150x100x1.6)' → '2P 6FT'
    Other items: strip prefixes/suffixes
    """
    n = name.strip()
    if not n:
        return n
    # For Column items: extract just "1P 6FT", "2P 8FT" etc.
    if 'column' in n.lower() or 'column' in subcat.lower():
        m = re.search(r'\b(\dP)\s+(\d+FT)\b', n, flags=re.IGNORECASE)
        if m:
            return m.group(1) + ' ' + m.group(2)
    # General shortening for non-column MMS items
    n = re.sub(r'^(?:GM\s+Bridge\s+)', '', n, flags=re.IGNORECASE)
    n = re.sub(r'^(?:Galvalume\s+)', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s*-\s*(?:SKU|ITEM|PROD)[-\s]?\w+\s*$', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s*\(\s*(?:SKU|ITEM|PROD)[-\s]?\w+\s*\)\s*$', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s*-\s*Solar\s*Square.*$', '', n, flags=re.IGNORECASE)
    return n.strip()[:80]


def shorten_cable_subcat(name):
    """Shorten cable subcategory names for cleaner dashboard display.
    e.g. 'Polycab 4 sqmm Cu DC Cable Red' → '4 sqmm Cu DC Cable'
    """
    n = name.strip()
    if not n:
        return n
    # Remove brand prefixes
    n = re.sub(r'^(?:Polycab|Havells|RR\s*Kabel|KEI|Finolex|Anchor)\s+', '', n, flags=re.IGNORECASE)
    # Remove colour suffixes like " Red", " Black", " Blue/Black"
    n = re.sub(r'\s+(?:Red|Black|Blue|Green|Yellow|White|Grey|Blue/Black|Red/Black)\s*$', '', n, flags=re.IGNORECASE)
    # Remove trailing codes
    n = re.sub(r'\s*-\s*(?:SKU|ITEM|PROD)[-\s]?\w+\s*$', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\s*-\s*Solar\s*Square.*$', '', n, flags=re.IGNORECASE)
    return n.strip()


# ── Build project map ─────────────────────────────────────────────────────────
print("\nReading data.csv.gz...")
project_map = {}
dn_metering = defaultdict(float)   # DN dump metering items per project
unmapped_cells = defaultdict(int)   # Track unmapped cells for warning
unmapped_cats  = defaultdict(int)   # Track categories not in CAT_KEY (diagnostic)
excluded_count = 0

# Sub-item aggregations for dashboard analysis
proj_inv_types = defaultdict(lambda: defaultdict(lambda: {'qty':0,'amt':0}))  # sse → inv_type → {qty, amt}
proj_mms_items = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {'qty':0,'amt':0,'uom':''})))  # sse → subcat → item_name → {qty, amt, uom}
proj_cable_items = defaultdict(lambda: defaultdict(lambda: {'qty':0,'amt':0,'cases':0}))  # sse → subcat → {qty, amt, cases}
proj_onm_amt = defaultdict(float)   # sse → total ONM amount
proj_qhse_amt = defaultdict(float)  # sse → total QHSE amount

with gzip.open('data.csv.gz', 'rt', encoding='utf-8', errors='replace') as f:
    reader = csv.DictReader(f)
    for i, row in enumerate(reader):
        if i % 100000 == 0: print(f"  {i:,} rows processed...")
        sse = row['SSE ID'].strip()
        if not sse: continue

        try: rev = float(row['Final Revenue Excl. GST']) if row['Final Revenue Excl. GST'].strip() else 0
        except: rev = 0
        try: kw = float(row['Project Size (kW)']) if row['Project Size (kW)'].strip() else 0
        except: kw = 0
        try: amt = float(row['amount']) if row['amount'].strip() else 0
        except: amt = 0
        try: qty = float(row['qty']) if row['qty'].strip() else 0
        except: qty = 0

        raw_cat   = row['item_category'].strip()
        item_code = row['item_code'].strip()
        item_name = row['item_name'].strip()
        item_subcat = row.get('item_subcategory', '').strip()
        parent_dn = row.get('parent', '').strip()
        uom = row.get('uom', '').strip()
        cat = resolve_cat(item_code, raw_cat, item_subcat)

        # Track parent-based ONM and QHSE amounts
        if parent_dn:
            pl = parent_dn.lower()
            if pl.startswith('onm'):
                proj_onm_amt[sse] += amt
            elif pl.startswith('qhse'):
                proj_qhse_amt[sse] += amt

        # Track DN dump metering items (from the Excel SUMIFS part of the formula)
        if is_metering_dn_item(item_name):
            dn_metering[sse] += amt

        if sse not in project_map:
            cell = row['Cell Name'].strip()
            cs   = CELL_CITY_STATE.get(cell)
            city = cs['c'] if cs else row['City'].strip()
            state= cs['s'] if cs else row['State'].strip()
            if cell and not cs and not city:
                unmapped_cells[cell] += 1
            d    = parse_date(row['Installation Completion Date'])
            offer_raw = row['Offer Type'].strip()
            offer= offer_raw.replace('GoodZero+','GoodZero')
            phase= row['Phase Connection'].strip()

            # ── Cohort assignment via QCD date from booking dump ──────────────
            qcd_rec = QCD_MAP.get(sse, {})
            qcd_dt  = qcd_rec.get('qcd', None)
            # Use offer from booking dump if available (more authoritative), else DN offer
            cohort_offer = qcd_rec.get('offer', '') or offer_raw
            cohort = assign_cohort(qcd_dt, cohort_offer)

            project_map[sse] = {
                'id':sse,'c':city,'s':state,'o':offer,'ph':phase,
                'kw':kw,'rev':round(rev,2),'dt':d.strftime('%Y-%m-%d') if d else '',
                'qcd': qcd_dt.strftime('%Y-%m-%d') if qcd_dt else '',
                'cohort': cohort,
                'mod':0,'inv':0,'prf':0,'cab':0,'ick':0,'con':0,'ear':0,'jbx':0,
                'tsh':0,'saf':0,'ica':0,'wel':0,'ssn':0,'ebo':0,'dlg':0,'mtr':0,'wkt':0,'lad':0,
                'mt':'','mq':0,'it':'','iq':0,
            }

        if cat == 'EXCLUDE':
            excluded_count += 1
            continue
        k2 = CAT_KEY.get(cat)
        if k2:
            project_map[sse][k2] = round(project_map[sse][k2] + amt, 2)
        elif cat:
            unmapped_cats[cat] += 1

        p = project_map[sse]
        if cat == 'Module' and item_name:
            if not p['mt']: p['mt'] = item_name; p['mq'] = qty
            elif p['mt'] == item_name: p['mq'] += qty
        if cat == 'Inverter' and item_name:
            if not p['it']:
                p['it'] = item_name; p['iq'] = qty
            # Detect inverter phase from DN item name (for NM rate lookup)
            if '_inv_phase' not in p or not p['_inv_phase']:
                detected = detect_inverter_phase(item_name)
                if detected:
                    p['_inv_phase'] = detected
            # Track inverter type for type-level analysis
            inv_type = detect_inverter_type(item_name)
            if inv_type:
                proj_inv_types[sse][inv_type]['qty'] += qty
                proj_inv_types[sse][inv_type]['amt'] += amt

        # Track MMS sub-items (Prefab MMS, Welded MMS, Tin Shed MMS)
        if cat in ('Prefab MMS', 'Welded MMS', 'Tin Shed MMS') and item_subcat:
            short_name = shorten_mms_item_name(item_name, item_subcat)
            proj_mms_items[sse][item_subcat][short_name]['qty'] += qty
            proj_mms_items[sse][item_subcat][short_name]['amt'] += amt
            proj_mms_items[sse][item_subcat][short_name]['uom'] = uom

        # Track cable sub-items
        if cat == 'Cables' and item_subcat:
            short_cable = shorten_cable_subcat(item_subcat)
            proj_cable_items[sse][short_cable]['qty'] += qty
            proj_cable_items[sse][short_cable]['amt'] += amt
            proj_cable_items[sse][short_cable]['cases'] += 1

print(f"Built {len(project_map):,} projects")
print(f"  Excluded rows (dongles, Safety Lifeline, Civil Work): {excluded_count:,}")

if unmapped_cells:
    print(f"\n⚠  WARNING: {len(unmapped_cells)} unmapped cell names (add to CELL_CITY_STATE):")
    for cell, cnt in sorted(unmapped_cells.items(), key=lambda x: -x[1]):
        print(f"    {cell}: {cnt} projects")

if unmapped_cats:
    print(f"\n⚠  WARNING: {len(unmapped_cats)} categories not in CAT_KEY (not counted in COGS):")
    for cat, cnt in sorted(unmapped_cats.items(), key=lambda x: -x[1])[:20]:
        print(f"    '{cat}': {cnt:,} rows")

# ── Backend metering injection (formula-based, dual-phase) ────────────────────
# metering = NM_rate(city, inv_phase) + GM_rate(city, sanction_phase) + DN_dump_items
# NM uses Inverter Phase (detected from DN item name via detect_inverter_phase)
# GM uses Sanction Phase (= Phase Connection from data.csv, stored in p['ph'])

month_metering = defaultdict(float)
no_rate_cities = defaultdict(int)
phase_mismatch_count = 0

for sse, p in project_map.items():
    inv_phase = p.get('_inv_phase', p['ph'])  # fallback to Phase Connection if no inverter detected
    sanction_phase = p['ph']                   # Phase Connection = Sanction Phase (100% match)
    if inv_phase != sanction_phase:
        phase_mismatch_count += 1

    backend = calc_metering_backend(p['c'], inv_phase, sanction_phase)
    dn = dn_metering.get(sse, 0)
    total_mtr = backend + dn

    if total_mtr > 0:
        p['mtr'] = round(p['mtr'] + total_mtr, 2)

    if p['dt']:
        mkey = p['dt'][:7]
        month_metering[mkey] += total_mtr

    if p['c'] and p['c'] not in NM_RATES and backend == 0:
        no_rate_cities[p['c']] += 1

print()
print(f"  Phase mismatches (inv_phase ≠ sanction_phase): {phase_mismatch_count}")
for mkey in sorted(month_metering):
    if month_metering[mkey] > 0:
        count = sum(1 for p in project_map.values() if p['dt'].startswith(mkey))
        print(f"  Metering {mkey}: ₹{month_metering[mkey]:,.0f} → {count} projects")

if no_rate_cities:
    print(f"\n⚠  Cities not in rate table (0 metering):")
    for c, cnt in sorted(no_rate_cities.items(), key=lambda x: -x[1]):
        print(f"    {c}: {cnt} projects")

# ── Compute final COGS ────────────────────────────────────────────────────────
projects = []
for sse, p in project_map.items():
    p.pop('_inv_phase', None)
    cogs = round(p['mod']+p['inv']+p['prf']+p['cab']+p['ick']+p['con']+p['ear']+
                 p['jbx']+p['tsh']+p['saf']+p['ica']+p['wel']+p['ssn']+p['ebo']+
                 p['dlg']+p['mtr']+p['wkt']+p['lad'], 2)

    out = {**p, 'cogs': cogs}

    # Add ONM and QHSE amounts from parent-based tracking
    onm_val = proj_onm_amt.get(sse, 0)
    qhse_val = proj_qhse_amt.get(sse, 0)
    if onm_val: out['onm'] = round(onm_val, 2)
    if qhse_val: out['qhs'] = round(qhse_val, 2)

    # Add inverter type breakdown
    ivt = proj_inv_types.get(sse)
    if ivt:
        out['ivt'] = {t: {'q': round(d['qty'],1), 'a': round(d['amt'],2)} for t, d in ivt.items()}

    # Add MMS sub-item breakdown: {subcat: {item_name: {q, a, u}}}
    mms = proj_mms_items.get(sse)
    if mms:
        out['msd'] = {}
        for subcat, items in mms.items():
            out['msd'][subcat] = {nm: {'q': round(d['qty'],2), 'a': round(d['amt'],2)} for nm, d in items.items()}

    # Add cable sub-item breakdown: {subcat: {q, a, n}}
    cab_items = proj_cable_items.get(sse)
    if cab_items:
        out['cbd'] = {sc: {'q': round(d['qty'],2), 'a': round(d['amt'],2), 'n': d['cases']} for sc, d in cab_items.items()}

    projects.append(out)

# ── Calculate monthly ONM & QHSE totals from raw data ──────────────────────────
# This ensures we capture ALL ONM/QHSE, not just those linked to specific projects
monthly_onm_qhse = defaultdict(lambda: {'onm': 0, 'qhs': 0})

with gzip.open('data.csv.gz', 'rt', encoding='utf-8', errors='replace') as f:
    reader = csv.DictReader(f)
    for row in reader:
        parent = row.get('parent', '').strip()
        posting_date_str = row.get('posting_date', '').strip()
        amount = float(row.get('amount', 0) or 0)

        if not posting_date_str or not parent:
            continue

        try:
            date_obj = datetime.strptime(posting_date_str, '%Y-%m-%d')
            month_key = f"{date_obj.year}-{date_obj.month:02d}"

            if parent.lower().startswith('onm'):
                monthly_onm_qhse[month_key]['onm'] += amount
            elif parent.lower().startswith('qhse'):
                monthly_onm_qhse[month_key]['qhs'] += amount
        except:
            pass

# Create metadata object with monthly totals
metadata = {
    'monthly_onm_qhse': {k: {'onm': round(v['onm'], 2), 'qhs': round(v['qhs'], 2)} for k, v in monthly_onm_qhse.items()}
}

# ── Write output ──────────────────────────────────────────────────────────────
output = {'_meta': metadata, 'projects': projects}
json_str = json.dumps(output, separators=(',',':'))

# Write compressed output using Python gzip (avoids shell gzip permission issues)
import shutil
with open('projects_temp.json', 'w', encoding='utf-8') as f:
    f.write(json_str)
with open('projects_temp.json', 'rb') as f_in, \
     gzip.open('projects.json.gz', 'wb', compresslevel=6) as f_out:
    shutil.copyfileobj(f_in, f_out)
try:
    os.remove('projects_temp.json')
except Exception:
    pass

raw_mb = len(json_str)/1e6
gz_mb  = os.path.getsize('projects.json.gz')/1e6
print(f"\nOutput: {len(projects):,} projects | JSON {raw_mb:.1f} MB → gz {gz_mb:.2f} MB")

# ── Quick verification ─────────────────────────────────────────────────────────
print("\n── Verification ──")

# Per-category totals for Jan/Feb 26 (diagnostic)
for mo, label in [(1, 'Jan 26'), (2, 'Feb 26'), (3, 'Mar 26')]:
    ps = [p for p in projects if p['dt'].startswith(f'2026-{mo:02d}')]
    if not ps: continue
    rev  = sum(p['rev']  for p in ps)
    cogs = sum(p['cogs'] for p in ps)
    mtr  = sum(p['mtr']  for p in ps)
    gm   = (rev-cogs)/rev*100 if rev else 0
    print(f"\n  {label}: {len(ps)} projects | Rev={rev/1e7:.2f}Cr | COGS={cogs/1e7:.2f}Cr | GM%={gm:.2f}%")
    # Breakdown by category
    for cat_name, key in sorted(CAT_KEY.items(), key=lambda x: -sum(p[x[1]] for p in ps)):
        total = sum(p[key] for p in ps)
        if total > 0:
            print(f"    {cat_name:25s}: ₹{total/1e7:.2f}Cr")

# Metering comparison with actuals
print("\n── Metering Accuracy ──")
for mo, label, actual_mtr in [(1, 'Jan 26', 5926077), (2, 'Feb 26', 5755707), (3, 'Mar 26', 7909163)]:
    ps = [p for p in projects if p['dt'].startswith(f'2026-{mo:02d}')]
    mtr = sum(p['mtr'] for p in ps)
    delta = mtr - actual_mtr
    pct   = delta / actual_mtr * 100 if actual_mtr else 0
    print(f"  {label}: Metering={mtr:,.0f} (actual {actual_mtr:,.0f}, delta {delta:+,.0f} = {pct:+.2f}%)")

# ── Cohort Coverage Report ────────────────────────────────────────────────────
print("\n── Cohort Assignment Coverage ──")
total_p = len(projects)
with_cohort = sum(1 for p in projects if p.get('cohort',''))
with_qcd    = sum(1 for p in projects if p.get('qcd',''))
print(f"  Projects with QCD date : {with_qcd:,} / {total_p:,} ({with_qcd/total_p*100:.1f}%)")
print(f"  Projects with cohort   : {with_cohort:,} / {total_p:,} ({with_cohort/total_p*100:.1f}%)")
if with_cohort < total_p:
    missing = total_p - with_cohort
    print(f"  \u26a0 {missing:,} projects have no cohort (missing from booking dump or QCD blank)")

print("\nDone.")
